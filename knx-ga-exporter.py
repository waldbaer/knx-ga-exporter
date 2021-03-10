#!/usr/bin/env python3

import logging
import sys
import traceback
import argparse
from openpyxl import load_workbook
import csv
from enum import Enum

# ---- Main ------------------------------------------------------------------------------------------------------------
class GroupAddress:
  def __init__(self, main, middle, sub, main_name, middle_name, sub_name, target_id, dpt, comment):
    self.main=main
    self.middle=middle
    self.sub=sub
    self.main_name=main_name
    self.middle_name=middle_name
    self.sub_name=sub_name
    self.target_id=target_id
    self.dpt=dpt
    self.comment=comment
  def __str__(self):
    return '{}/{}/{} {}|{}|{}'.format(self.main, self.middle, self.sub, self.main_name, self.middle_name, self.sub_name)

def main():
  args = ParseCommandLineArguments()
  logging.basicConfig(level=logging.DEBUG)

  wb = LoadWorkbook(args.input_file)
  gas = ParseGroupAddresses(wb, args)
  ExportCsv(args, gas)

  logging.info("Statistics: #GA: {}".format(len(gas)))
  logging.info('done.')

class CsvFormat(Enum):
  format_3_1 = '3/3' # Main- Middle- Sub- Name/Main- Middle- Sub-Address
  format_1_1 = '1/1' # Name / Address

  def __str__(self):
      return self.value

class CsvSeparator(Enum):
  tabulator = "tabulator"
  comma = "comma"
  semicolon = "semicolon"

  def __str__(self):
      return self.value

def ParseCommandLineArguments():
  parser = argparse.ArgumentParser(description='KNX group address exporter.')

  parser.add_argument('-i', '--input', dest='input_file', required=True, help='Path to XSLX file to be parsed.')
  parser.add_argument('-o', '--output', dest='output_file', required=False, default='knx-ga-addresses.csv', help='Path of exported CSV file.')
  parser.add_argument('--output-encoding', dest='output_file_encoding', required=False, default='iso-8859-1', help='Output file encoding')

  parser.add_argument('--csv-format', default=CsvFormat.format_1_1, type=CsvFormat, choices=list(CsvFormat), help='CSV output format.')
  parser.add_argument('--csv-separator', default=CsvSeparator.tabulator, type=CsvSeparator, choices=list(CsvSeparator), help='CSV separator.')

  parser.add_argument('--ga-sheet-name', default='KNX Group Addresses', help='Name of XLSX sheet containing the KNX group addresses')
  parser.add_argument('--ga-sheet-first-row', default=8, help='First row containing GAs')
  parser.add_argument('--ga-sheet-last-column', default=10, help='Last column')

  parser.add_argument('--ga-sheet-main-ID-column', default=0, help='Column containing main ID of KNX GA')
  parser.add_argument('--ga-sheet-middle-ID-column', default=2, help='Column containing middle ID of KNX GA')
  parser.add_argument('--ga-sheet-sub-ID-column', default=4, help='Column containing sub ID of KNX GA')

  parser.add_argument('--ga-sheet-main-name-column', default=1, help='Column containing main name of KNX GA')
  parser.add_argument('--ga-sheet-middle-name-column', default=3, help='Column containing middle name of KNX GA')
  parser.add_argument('--ga-sheet-sub-name-column', default=8, help='Column containing sub name of KNX GA')

  parser.add_argument('--ga-sheet-dpt-column', default=5, help='Column containing datapoint type of KNX GA')
  parser.add_argument('--ga-sheet-target-ID-column', default=6, help='Column containing target ID KNX GA')
  parser.add_argument('--ga-sheet-compiled-GA-column', default=7, help='Column containing full accumulated GA')
  parser.add_argument('--ga-sheet-comment', default=9, help='Column containing GA comment')

  args = parser.parse_args()
  return args

def LoadWorkbook(input_file):
  logging.info("Loading XLSX input file '{}'".format(input_file))
  wb = load_workbook(filename=input_file, read_only=True, data_only=True)
  return wb

def ParseGroupAddresses(wb, args):
  gas = []
  ws = wb[args.ga_sheet_name]
  for row in ws.iter_rows(min_row=args.ga_sheet_first_row, max_col=args.ga_sheet_last_column):
    target_id = row[args.ga_sheet_target_ID_column].value

    group_main = row[args.ga_sheet_main_ID_column].value
    group_middle = row[args.ga_sheet_middle_ID_column].value
    group_sub = row[args.ga_sheet_sub_ID_column].value

    group_main_name = row[args.ga_sheet_main_name_column].value
    group_middle_name = row[args.ga_sheet_middle_name_column].value
    group_sub_name = row[args.ga_sheet_sub_name_column].value

    dpt = row[args.ga_sheet_dpt_column].value
    compiled_ga = row[args.ga_sheet_compiled_GA_column].value
    comment = row[args.ga_sheet_comment].value

    # Skip invalid / incomplete GAs
    if(dpt == None or dpt == 0 or compiled_ga == None or compiled_ga == 0):
      continue

    ga = GroupAddress(group_main, group_middle, group_sub, group_main_name, group_middle_name, group_sub_name, target_id, dpt, comment)
    logging.info("Parsed GA: %s" % (str(ga)))
    gas.append(ga)

  return gas

def ExportCsv(args, gas):
  csv_separators = {"tabulator" : '\t',
                    "comma" : ',',
                    "semicolon" : ';'}
  exporter_functions = {"1/1" : ExportCsvFormat1_1,
                        "3/3" : ExportCsvFormat3_3}

  csv_separator = csv_separators[str(args.csv_separator)]
  export_function = exporter_functions[str(args.csv_format)]
  export_function(args.output_file, args.output_file_encoding, csv_separator, gas)

def ExportCsvFormat1_1(output_file, output_file_encoding, csv_separator, gas):
  logging.info("Exporting group addresses into CSV file '{}'. Format: 1/1, separator:{}".format(output_file, csv_separator))
  with open(output_file, 'w', newline='', encoding=output_file_encoding) as csvfile:
    writer = csv.writer(csvfile, delimiter=csv_separator, quoting=csv.QUOTE_ALL)

    # write headline
    writer.writerow(["Group name", "Address", "Central", "Unfiltered", "Description", "DatapointType", "Security"])

    main_group_ids = {ga.main for ga in gas}
    for main_group_id in main_group_ids:
      filtered_main_gas = list(filter(lambda ga: ga.main == main_group_id, gas))
      main_group_name = filtered_main_gas[0].main_name

      logging.info("Exporting main group {}: {}".format(main_group_id, main_group_name))
      writer.writerow([main_group_name, '{}/-/-'.format(main_group_id)] + [''] * 4 + ['Auto'])

      middle_group_ids = {ga.middle for ga in filtered_main_gas}
      for middle_group_id in middle_group_ids:
        filtered_middle_gas = list(filter(lambda ga: ga.middle == middle_group_id, filtered_main_gas))
        middle_group_name = filtered_middle_gas[0].middle_name

        logging.info("Exporting middle group {}/{}: {}".format(main_group_id, middle_group_id, middle_group_name))
        writer.writerow([middle_group_name, '{}/{}/-'.format(main_group_id, middle_group_id)] + [''] * 4 + ['Auto'])

        for sub_ga in filtered_middle_gas:
          logging.info("Exporting sub group: {}".format(sub_ga))
          ga_name = FormatGaName(sub_ga)
          ga_description = FormatGaDescription(sub_ga)
          writer.writerow([ga_name, '{}/{}/{}'.format(sub_ga.main, sub_ga.middle, sub_ga.sub), '', '', ga_description, sub_ga.dpt, 'Auto'])

def ExportCsvFormat3_3(output_file, output_file_encoding, csv_separator, gas):
  logging.info("Exporting group addresses into CSV file '{}'. Format: 3/3, separator:{}".format(output_file, csv_separator))
  with open(output_file, 'w', newline='', encoding=output_file_encoding) as csvfile:
    writer = csv.writer(csvfile, delimiter=csv_separator, quoting=csv.QUOTE_ALL)

    # write headline
    writer.writerow(['Main', 'Middle', 'Sub', 'Main', 'Middle', 'Sub', 'Central', 'Unfiltered', 'Description', 'DatapointType','Security'])

    main_group_ids = {ga.main for ga in gas}
    for main_group_id in main_group_ids:
      filtered_main_gas = list(filter(lambda ga: ga.main == main_group_id, gas))
      main_group_name = filtered_main_gas[0].main_name

      logging.info("Exporting main group {}: {}".format(main_group_id, main_group_name))
      writer.writerow([main_group_name, '', '', main_group_id] + [''] * 6 + ['Auto'])

      middle_group_ids = {ga.middle for ga in filtered_main_gas}
      for middle_group_id in middle_group_ids:
        filtered_middle_gas = list(filter(lambda ga: ga.middle == middle_group_id, filtered_main_gas))
        middle_group_name = filtered_middle_gas[0].middle_name

        logging.info("Exporting middle group {}/{}: {}".format(main_group_id, middle_group_id, middle_group_name))
        writer.writerow(['', middle_group_name, '', main_group_id, middle_group_id] + [''] * 5 + ['Auto'])

        for sub_ga in filtered_middle_gas:
          logging.info("Exporting sub group: {}".format(sub_ga))
          ga_name = FormatGaName(sub_ga)
          ga_description = FormatGaDescription(sub_ga)
          writer.writerow(['', '', ga_name, sub_ga.main, sub_ga.middle, sub_ga.sub, '', '', ga_description, sub_ga.dpt, 'Auto'])

def FormatGaName(ga):
  ga_name = ga.sub_name
  if(ga.target_id != None and ga.target_id != 0):
    ga_name = ga.target_id + " - " + ga_name
    ga_name = ga_name.replace('\n', '_')
  return ga_name

def FormatGaDescription(ga):
  description = ga.comment
  # Replace line-breaks
  if(description != None):
    description = description.replace('\n', '|')
  return description

# ---- Entrypoint ------------------------------------------------------------------------------------------------------
if __name__ == '__main__':
  try:
    main()
  except SystemExit:
    sys.exit(1)
  except BaseException:
    logging.error("Any error has occured! Traceback:\r\n" + traceback.format_exc())
    sys.exit(1)
  sys.exit(0)
