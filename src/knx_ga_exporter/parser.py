"""Workbook parser."""

# ---- Imports ---------------------------------------------------------------------------------------------------------
import logging

import openpyxl
import openpyxl.workbook

from knx_ga_exporter.group_address import GroupAddress

# ---- Functions -------------------------------------------------------------------------------------------------------


def load_workbook(input_file: str) -> openpyxl.workbook:
    """Load XLX workbook.

    Args:
        input_file (str): Path of input file.

    Returns:
        Loaded workbook
    """
    logging.info("Loading XLSX input file '%s'", input_file)
    wb = openpyxl.load_workbook(filename=input_file, read_only=True, data_only=True)
    return wb


def parse_group_addresses(wb: openpyxl.workbook, layout_config: dict) -> list[GroupAddress]:
    """Parse the group addresses.

    Arguments:
        wb: Workbook
        layout_config: Workbook layout configuration

    Returns:
        Parsed KNX group addresses
    """
    gas = []
    ws = wb[layout_config.sheet_name]
    for row in ws.iter_rows(min_row=layout_config.first_row, max_col=layout_config.last_column):
        target_id = row[layout_config.target_ID_column].value

        group_main = row[layout_config.main_ID_column].value
        group_middle = row[layout_config.middle_ID_column].value
        group_sub = row[layout_config.sub_ID_column].value

        group_main_name = row[layout_config.main_name_column].value
        group_middle_name = row[layout_config.middle_name_column].value
        group_sub_name = row[layout_config.sub_name_column].value

        dpt = row[layout_config.dpt_column].value
        comment = row[layout_config.comment_column].value

        # Skip invalid / incomplete GAs
        if dpt is None or dpt == 0 or dpt is None:
            continue

        ga = GroupAddress(
            group_main,
            group_middle,
            group_sub,
            group_main_name,
            group_middle_name,
            group_sub_name,
            target_id,
            dpt,
            comment,
        )
        logging.debug("Parsed GA: %s", ga)
        gas.append(ga)

    return gas
